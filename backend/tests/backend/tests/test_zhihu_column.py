# -*- coding: utf-8 -*-
"""知乎专栏下载模块单元测试（纯 Python，无需 pytest）"""
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import zhihu_column

PASS = 0
FAIL = 0


def check(name: str, cond: bool, detail: str = ""):
    global PASS, FAIL
    if cond:
        PASS += 1
        print(f"  OK {name}")
    else:
        FAIL += 1
        print(f"  FAIL {name} {detail}")


def test_normalize_column_input():
    check(
        "裸 ID",
        zhihu_column.normalize_column_input("c_1020247688083775488") == "c_1020247688083775488",
    )
    check(
        "专栏 URL",
        zhihu_column.normalize_column_input("https://www.zhihu.com/column/c_1020247688083775488")
        == "c_1020247688083775488",
    )
    check(
        "zhuanlan URL",
        zhihu_column.normalize_column_input("https://zhuanlan.zhihu.com/wangzhenotes")
        == "wangzhenotes",
    )
    try:
        zhihu_column.normalize_column_input("")
        check("空输入报错", False)
    except ValueError:
        check("空输入报错", True)

    # 个人主页应走 resolve_column_input（自动解析同名专栏），normalize 本身保持严格
    try:
        zhihu_column.normalize_column_input("https://www.zhihu.com/people/pydatalysis")
        check("个人主页 URL 不被 normalize 接受", False)
    except ValueError:
        check("个人主页 URL 不被 normalize 接受", True)


def test_resolve_column_input():
    col, note = zhihu_column.resolve_column_input("pydatalysis")
    check("裸 token 解析", col == "pydatalysis" and note == "")

    col, note = zhihu_column.resolve_column_input("https://zhuanlan.zhihu.com/pydatalysis")
    check("专栏链接解析", col == "pydatalysis" and note == "")

    col, note = zhihu_column.resolve_column_input("https://www.zhihu.com/column/pydatalysis")
    check("column 链接解析", col == "pydatalysis" and note == "")

    try:
        zhihu_column.resolve_column_input("")
        check("空输入报错", False)
    except ValueError:
        check("空输入报错", True)


def test_html_to_markdown():
    html = """
    <div class="RichContent-inner">
      <h2>标题</h2>
      <p>第一段 <strong>加粗</strong> 和 <em>斜体</em></p>
      <figure><img data-actualsrc="https://pic.zhimg.com/v2-abc.jpg" alt="图片描述"></figure>
      <img class="ztext-math" data-tex="E=mc^2">
      <ul><li>项目一</li><li>项目二</li></ul>
      <blockquote>引用内容</blockquote>
      <div class="video-box"><a href="https://www.zhihu.com/zvideo/123">视频标题</a></div>
    </div>
    """
    md = zhihu_column.html_to_markdown(html, "https://zhuanlan.zhihu.com/p/123")
    check("正文文本", "第一段" in md, md[:200])
    check("加粗", "**加粗**" in md, md)
    check("图片", "![图片描述](https://pic.zhimg.com/v2-abc.jpg)" in md, md)
    check("公式", "E=mc^2" in md, md)
    check("列表", "- 项目一" in md, md)
    check("引用", "> 引用内容" in md, md)
    check("视频链接", "zvideo/123" in md, md)


def test_build_markdown():
    item = {
        "type": "article",
        "title": "测试文章",
        "url": "https://zhuanlan.zhihu.com/p/1",
        "author": "张三",
        "created": "2024-01-01 10:00:00",
        "updated": "2024-01-02 10:00:00",
        "excerpt": "摘要内容",
        "comments": 5,
        "likes": 10,
    }
    md = zhihu_column.build_markdown(item, "正文内容")
    check("标题", "# 测试文章" in md)
    check("作者", "作者：张三" in md)
    check("链接", "https://zhuanlan.zhihu.com/p/1" in md)
    check("互动", "赞同：10 / 评论：5" in md)
    check("正文", "正文内容" in md)


def test_generate_excel():
    rows = [
        {"type": "article", "title": "文章一", "url": "https://a", "created": "2024-01-01",
         "updated": "", "excerpt": "简介", "comments": 3, "likes": 8},
        {"type": "answer", "title": "回答一", "url": "https://b", "created": "2024-01-02",
         "updated": "", "excerpt": "简介2", "comments": 1, "likes": 2},
    ]
    with tempfile.TemporaryDirectory() as td:
        path = zhihu_column.generate_excel(rows, Path(td) / "test.xlsx")
        check("Excel 文件生成", path.exists())
        from openpyxl import load_workbook
        wb = load_workbook(path)
        ws = wb.active
        check("表头", ws.cell(row=1, column=1).value == "序号")
        check("第一行数据", ws.cell(row=2, column=3).value == "文章一")
        check("赞同数", ws.cell(row=3, column=9).value == 2)


def main():
    print("== 知乎专栏模块测试 ==")
    test_normalize_column_input()
    test_resolve_column_input()
    test_html_to_markdown()
    test_build_markdown()
    test_generate_excel()
    print(f"\n结果: {PASS} 通过, {FAIL} 失败")
    sys.exit(1 if FAIL else 0)


if __name__ == "__main__":
    main()
