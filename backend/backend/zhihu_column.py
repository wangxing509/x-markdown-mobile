# -*- coding: utf-8 -*-
"""
知乎专栏批量下载模块

输入专栏 ID（或链接），自动抓取该专栏下的文章 / 回答 / 视频：
- 每篇内容转换为 Markdown 保存到本地
- 生成包含文章数据的 Excel 表格（类型 / 标题 / 链接 / 时间 / 简介 / 评论 / 赞同）
- 可选下载视频文件（默认仅保留视频链接）

依赖：httpx / beautifulsoup4 / markdownify / openpyxl
"""
import json
import re
import threading
import time
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

import httpx
from bs4 import BeautifulSoup

from config import USER_DIR, HTTP_PROXY

# ==================== 常量 ====================

ZHIHU_UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/127.0.0.0 Safari/537.36"
)

COOKIE_PATH = Path.home() / ".zhihu_cookie.json"
XMD_COOKIE_PATH = USER_DIR / "zhihu_cookie.json"

DEFAULT_OUTPUT_DIR = USER_DIR / "zhihu"

PAGE_SIZE = 10  # zhuanlan items 接口实际单页上限
REQUEST_TIMEOUT = 25

_JOBS: dict[str, dict] = {}
_JOBS_LOCK = threading.Lock()

_COLUMN_RE = re.compile(r"zhuanlan\.zhihu\.com/([A-Za-z0-9_-]+)")
_COLUMN_URL_RE = re.compile(r"zhihu\.com/column/([A-Za-z0-9_-]+)")
_PEOPLE_RE = re.compile(r"zhihu\.com/people/([A-Za-z0-9_-]+)")


# ==================== 工具函数 ====================

def _now() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _ensure_utf8_io():
    """确保控制台输出 UTF-8（Windows 兼容）"""
    import io
    import sys

    if sys.stdout and sys.stdout.encoding != "utf-8":
        try:
            sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
        except Exception:
            pass
    if sys.stderr and sys.stderr.encoding != "utf-8":
        try:
            sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")
        except Exception:
            pass


def _slugify(text: str, max_len: int = 60) -> str:
    """生成安全文件名（保留中文、字母、数字）"""
    slug = re.sub(r'[\\/:*?"<>|\r\n\t]+', "_", text.strip())
    slug = re.sub(r"\s+", " ", slug).strip()
    slug = re.sub(r"[_ ]{2,}", "_", slug).strip("_")
    return slug[:max_len] or "untitled"


def normalize_column_input(value: str) -> str:
    """从输入（ID 或 URL）中提取专栏 ID"""
    raw = (value or "").strip()
    if not raw:
        raise ValueError("请输入知乎专栏 ID 或链接")
    m = _COLUMN_RE.search(raw)
    if not m:
        m = _COLUMN_URL_RE.search(raw)
    col_id = m.group(1) if m else raw
    col_id = col_id.rstrip("/")
    if not re.fullmatch(r"[A-Za-z0-9_-]+", col_id):
        raise ValueError(f"无法识别的专栏 ID: {col_id}")
    return col_id


def _column_exists(col_id: str) -> bool:
    """检查专栏是否存在（用于个人主页 → 同名专栏解析）"""
    try:
        fetch_column_info(col_id)
        return True
    except Exception:
        return False


def resolve_column_input(value: str) -> tuple[str, str]:
    """解析输入为专栏 ID；个人主页链接会自动尝试解析为同名专栏

    返回: (column_id, 解析说明)
    """
    raw = (value or "").strip()
    if not raw:
        raise ValueError("请输入知乎专栏 ID 或链接")

    m = _COLUMN_RE.search(raw)
    if m:
        return m.group(1), ""
    m = _COLUMN_URL_RE.search(raw)
    if m:
        return m.group(1), ""

    m = _PEOPLE_RE.search(raw)
    if m:
        token = m.group(1)
        if _column_exists(token):
            return token, f"已从个人主页自动解析到同名专栏「{token}」"
        raise ValueError(
            f"这是个人主页（{raw}），不是专栏，且未找到同名专栏。"
            f"请改为专栏链接，例如 https://zhuanlan.zhihu.com/{token}"
        )

    col_id = raw.rstrip("/")
    if not re.fullmatch(r"[A-Za-z0-9_-]+", col_id):
        raise ValueError(f"无法识别的专栏 ID: {col_id}")
    return col_id, ""


def load_cookie() -> dict:
    """读取知乎 Cookie（优先 ~/.zhihu_cookie.json，其次 ~/.xmarkdown/zhihu_cookie.json）"""
    for path in (COOKIE_PATH, XMD_COOKIE_PATH):
        if path.exists():
            try:
                data = json.loads(path.read_text(encoding="utf-8"))
                if isinstance(data, dict) and data.get("z_c0"):
                    return data
            except Exception as e:
                print(f"  [知乎 Cookie] 读取 {path} 失败: {e}")
    return {}


def save_cookie(z_c0: str, d_c0: str = "") -> None:
    """保存知乎 Cookie 到 ~/.zhihu_cookie.json"""
    z_c0 = (z_c0 or "").strip()
    if not z_c0:
        raise ValueError("z_c0 不能为空")
    data = load_cookie()
    data["z_c0"] = z_c0
    if d_c0:
        data["d_c0"] = d_c0
    data["saved_at"] = datetime.now().isoformat(timespec="seconds")
    COOKIE_PATH.write_text(
        json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def cookie_status() -> dict:
    data = load_cookie()
    z_c0 = data.get("z_c0", "")
    return {
        "hasCookie": bool(z_c0),
        "savedAt": data.get("saved_at") or "",
        "zC0Masked": (z_c0[:8] + "..." + z_c0[-6:]) if len(z_c0) > 16 else "",
    }


def _build_headers(col_id: str = "") -> dict:
    data = load_cookie()
    cookie_parts = []
    if data.get("z_c0"):
        cookie_parts.append(f"z_c0={data['z_c0']}")
    if data.get("d_c0"):
        cookie_parts.append(f"d_c0={data['d_c0']}")
    referer = f"https://www.zhihu.com/column/{col_id}" if col_id else "https://www.zhihu.com/"
    return {
        "User-Agent": ZHIHU_UA,
        "Accept": "application/json, text/plain, */*",
        "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
        "Referer": referer,
        "Origin": "https://www.zhihu.com",
        "x-requested-with": "fetch",
        "Cookie": "; ".join(cookie_parts),
    }


def _new_client(col_id: str = "") -> httpx.Client:
    return httpx.Client(
        headers=_build_headers(col_id),
        timeout=httpx.Timeout(REQUEST_TIMEOUT, connect=8),
        follow_redirects=True,
        proxy=HTTP_PROXY if HTTP_PROXY else None,
    )


# ==================== 知乎 API 抓取 ====================

def fetch_column_info(col_id: str) -> dict:
    """获取专栏基本信息"""
    url = (
        f"https://www.zhihu.com/api/v4/columns/{col_id}"
        f"?include=name,description,image_url,items_count,articles_count"
    )
    client = _new_client(col_id)
    try:
        resp = client.get(url)
        if resp.status_code != 200:
            raise RuntimeError(f"专栏信息接口 HTTP {resp.status_code}")
        data = resp.json()
        return {
            "columnId": col_id,
            "title": data.get("title") or col_id,
            "description": data.get("description") or "",
            "imageUrl": data.get("image_url") or "",
            "itemsCount": data.get("items_count") or data.get("articles_count") or 0,
            "author": (data.get("author") or {}).get("name", "") if isinstance(data.get("author"), dict) else "",
        }
    finally:
        client.close()


def inspect_column(value: str) -> dict:
    """解析输入并返回专栏信息（供界面在下载前展示作者与总条数）"""
    col_id, resolved_from = resolve_column_input(value)
    info = fetch_column_info(col_id)
    info["columnId"] = col_id
    info["resolvedFrom"] = resolved_from
    return info


def _items_endpoints(col_id: str) -> list[str]:
    """候选条目接口：v4 items → zhuanlan items → v4 articles（仅文章）"""
    return [
        f"https://www.zhihu.com/api/v4/columns/{col_id}/items",
        f"https://zhuanlan.zhihu.com/api/columns/{col_id}/items",
        f"https://www.zhihu.com/api/v4/columns/{col_id}/articles",
    ]


def _fetch_items_page(client: httpx.Client, url: str, offset: int) -> tuple[list[dict], bool]:
    """抓取一页条目，返回 (items, is_end)"""
    resp = client.get(url, params={"limit": 20, "offset": offset})
    if resp.status_code != 200:
        raise RuntimeError(f"条目接口 HTTP {resp.status_code}")
    data = resp.json()
    items = data.get("data") or []
    paging = data.get("paging") or {}
    return items, bool(paging.get("is_end"))


def fetch_column_items(col_id: str, max_items: int = 0, log: Optional[callable] = None) -> list[dict]:
    """分页抓取专栏全部条目（文章/回答/视频）"""
    client = _new_client(col_id)
    try:
        last_error: Optional[Exception] = None
        for endpoint in _items_endpoints(col_id):
            try:
                offset = 0
                raw_items: list[dict] = []
                while True:
                    page, is_end = _fetch_items_page(client, endpoint, offset)
                    if not page:
                        break
                    raw_items.extend(page)
                    offset += len(page)
                    if max_items and len(raw_items) >= max_items:
                        raw_items = raw_items[:max_items]
                        break
                    if is_end:
                        break
                    if log:
                        log(f"  已扫描 {len(raw_items)} 条...")
                return raw_items
            except Exception as e:
                last_error = e
                if log:
                    log(f"  接口 {endpoint} 失败，尝试备用接口: {e}")
                continue
        raise RuntimeError(f"所有条目接口均失败: {last_error}")
    finally:
        client.close()


def _extract_item(item: dict) -> dict:
    """规范化单条专栏条目"""
    item_type = item.get("type") or "article"
    author = item.get("author") or {}
    author_name = author.get("name", "") if isinstance(author, dict) else ""

    if item_type == "article":
        title = (item.get("title") or "未命名文章").strip()
        url = f"https://zhuanlan.zhihu.com/p/{item.get('id')}"
    elif item_type == "answer":
        question = item.get("question") or {}
        title = (question.get("title") or "知乎回答").strip()
        qid = question.get("id", "")
        url = f"https://www.zhihu.com/question/{qid}/answer/{item.get('id')}"
    elif item_type in ("video", "zvideo"):
        title = (item.get("title") or "知乎视频").strip()
        url = f"https://www.zhihu.com/zvideo/{item.get('id')}"
    else:
        title = (item.get("title") or f"知乎{item_type}").strip()
        url = item.get("url") or f"https://www.zhihu.com/{item_type}/{item.get('id')}"

    created = item.get("created_time") or item.get("created") or 0
    updated = item.get("updated_time") or item.get("updated") or created

    def _fmt_ts(ts):
        try:
            return datetime.fromtimestamp(int(ts)).strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            return ""

    return {
        "type": item_type,
        "id": str(item.get("id") or ""),
        "title": title,
        "url": url,
        "created": _fmt_ts(created),
        "updated": _fmt_ts(updated),
        "excerpt": (item.get("excerpt") or item.get("excerpt_title") or "").strip(),
        "comments": int(item.get("comment_count") or 0),
        "likes": int(item.get("voteup_count") or 0),
        "content": item.get("content") or "",
        "contentTruncated": bool(item.get("content_need_truncated")),
        "author": author_name,
        "video": item.get("video") if isinstance(item.get("video"), dict) else None,
        "raw": item,
    }


def _fetch_item_content(item: dict) -> str:
    """获取完整正文；条目自带 content 则直接使用，否则抓取 HTML 兜底"""
    content = item.get("content") or ""
    if content and not item.get("contentTruncated"):
        return content
    if content and len(content) > 500:
        return content

    client = _new_client()
    try:
        # 1) 回答/文章详情 API（实测 answer API 可用，article API 视风控而定）
        if item["type"] in ("article", "answer") and item.get("id"):
            kind = "answers" if item["type"] == "answer" else "articles"
            url = f"https://www.zhihu.com/api/v4/{kind}/{item['id']}?include=content"
            resp = client.get(url)
            if resp.status_code == 200:
                data = resp.json()
                c = data.get("content") or ""
                if c and len(c) > 100:
                    return c

        # 2) HTML 页面兜底
        page_url = item["url"]
        resp = client.get(page_url)
        if resp.status_code != 200:
            return content
        soup = BeautifulSoup(resp.text, "lxml")
        node = (
            soup.find("div", class_="RichContent-inner")
            or soup.find("div", class_="Post-RichTextContainer")
            or soup.find("div", class_="rich-content")
            or soup.find("div", class_="ztext")
            or soup.find("article")
            or soup.find("main")
        )
        return str(node) if node else content
    except Exception as e:
        print(f"  [知乎正文] HTML 兜底失败 {page_url}: {e}")
        return content
    finally:
        client.close()


# ==================== HTML → Markdown ====================

def html_to_markdown(html: str, source_url: str = "") -> str:
    """将知乎正文 HTML 转为 Markdown（处理懒加载图片、数学公式、视频卡片）"""
    if not html or not html.strip():
        return ""
    soup = BeautifulSoup(html, "lxml")

    # 1) 视频卡片 → 链接（在删除无关标签前处理）
    for box in soup.find_all(["div", "figure"], class_=lambda c: c and "video" in str(c).lower()):
        a = box.find("a")
        href = ""
        if a and a.get("href"):
            href = a["href"]
        elif box.get("data-video-url"):
            href = box["data-video-url"]
        text = (box.get_text(" ", strip=True) or "查看视频").strip()
        title_tag = box.find(class_=lambda c: c and "title" in str(c).lower())
        if title_tag:
            text = title_tag.get_text(strip=True)
        if href:
            box.replace_with(soup.new_string(f"\n\n[▶ {text}]({href})\n\n"))
        else:
            box.replace_with(soup.new_string(f"\n\n[▶ {text}]({source_url})\n\n"))

    # 2) 数学公式（知乎以 data-tex 或 img alt 内嵌 TeX）
    for tag in soup.find_all(["img", "span"], class_=lambda c: c and "math" in str(c).lower()):
        tex = tag.get("data-tex") or tag.get("alt") or ""
        if tex and ("$" not in tex):
            tag.replace_with(soup.new_string(f"${tex}$"))

    # 3) 懒加载图片：优先 data-actualsrc / data-original / data-src
    for img in soup.find_all("img"):
        src = (
            img.get("data-actualsrc")
            or img.get("data-original")
            or img.get("data-src")
            or img.get("src")
            or ""
        )
        if not src or src.startswith("data:") or "emoji" in src.lower():
            img.decompose()
            continue
        alt = img.get("alt") or ""
        img.replace_with(soup.new_string(f"![{alt}]({src})"))

    # 4) 清理无关元素
    for tag in soup.find_all(
        ["script", "style", "noscript", "iframe", "svg", "form", "button", "input"]
    ):
        tag.decompose()
    for tag in soup.find_all(class_=lambda c: c and any(
        x in str(c).lower() for x in ["sidebar", "ad-", "banner", "comment-list", "recommend", "share", "copyright"]
    )):
        tag.decompose()

    # 5) markdownify 转换
    try:
        from markdownify import markdownify as md
        markdown = md(str(soup), heading_style="ATX", bullets="-")
    except ImportError:
        text = soup.get_text("\n", strip=True)
        markdown = text

    # 6) 后处理：压缩空行、去掉首尾空白
    lines = []
    prev_blank = False
    for line in markdown.split("\n"):
        blank = line.strip() == ""
        if blank and prev_blank:
            continue
        lines.append(line.rstrip())
        prev_blank = blank
    return "\n".join(lines).strip()


def build_markdown(item: dict, content_md: str) -> str:
    """组装单篇 Markdown 文件内容"""
    parts = [f"# {item['title']}"]
    meta = []
    if item.get("author"):
        meta.append(f"- 作者：{item['author']}")
    meta.append(f"- 类型：{item['type']}")
    meta.append(f"- 链接：{item['url']}")
    if item.get("created"):
        meta.append(f"- 创建时间：{item['created']}")
    if item.get("updated"):
        meta.append(f"- 更新时间：{item['updated']}")
    meta.append(f"- 赞同：{item['likes']} / 评论：{item['comments']}")
    parts.append("\n".join(meta))
    if item.get("excerpt"):
        parts.append(f"> {item['excerpt']}")
    parts.append(content_md or "> （正文获取失败，请打开原链接查看）")
    parts.append("---")
    parts.append(f"> 来源：{item['url']}")
    return "\n\n".join(parts)


# ==================== 视频下载 ====================

def _find_video_url(video: Optional[dict]) -> str:
    """从视频对象提取 mp4 直链"""
    if not video:
        return ""
    playlist = video.get("playlist") or {}
    for quality in ("hd", "sd", "ld", "high", "medium", "low", "default"):
        entry = playlist.get(quality)
        if isinstance(entry, dict):
            u = entry.get("play_url") or entry.get("url") or ""
            if u:
                return u
        elif isinstance(entry, str) and entry:
            return entry
    return video.get("play_url") or video.get("video_url") or ""


def download_video(url: str, dest: Path) -> bool:
    """下载视频文件到 dest"""
    if not url:
        return False
    dest.parent.mkdir(parents=True, exist_ok=True)
    headers = {"User-Agent": ZHIHU_UA, "Referer": "https://www.zhihu.com/"}
    try:
        with httpx.Client(
            headers=headers,
            timeout=httpx.Timeout(60, connect=10),
            follow_redirects=True,
            proxy=HTTP_PROXY if HTTP_PROXY else None,
        ) as client:
            with client.stream("GET", url) as resp:
                if resp.status_code != 200:
                    return False
                with open(dest, "wb") as f:
                    for chunk in resp.iter_bytes(1024 * 256):
                        f.write(chunk)
        return dest.stat().st_size > 1024
    except Exception as e:
        print(f"  [视频下载] 失败 {url}: {e}")
        return False


# ==================== Excel 导出 ====================

def generate_excel(rows: list[dict], path: Path) -> Path:
    """生成文章数据 Excel（类型/标题/链接/时间/简介/评论/赞同）"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        raise RuntimeError(
            "缺少 openpyxl 依赖，请先执行: pip install openpyxl"
        ) from e

    headers = ["序号", "类型", "标题", "链接", "创建时间", "更新时间", "简介", "评论数", "赞同数"]
    wb = Workbook()
    ws = wb.active
    ws.title = "文章数据"
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for i, r in enumerate(rows, 1):
        ws.append([
            i,
            r.get("type", ""),
            r.get("title", ""),
            r.get("url", ""),
            r.get("created", ""),
            r.get("updated", ""),
            r.get("excerpt", ""),
            int(r.get("comments", 0) or 0),
            int(r.get("likes", 0) or 0),
        ])
        ws.cell(row=i + 1, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=i + 1, column=8).alignment = Alignment(horizontal="center")
        ws.cell(row=i + 1, column=9).alignment = Alignment(horizontal="center")

    widths = [6, 10, 48, 56, 20, 20, 60, 10, 10]
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    ws.freeze_panes = "A2"
    wb.save(path)
    return path


# ==================== 后台任务 ====================

def _log(job: dict, msg: str):
    job["logs"].append(f"[{_now()}] {msg}")
    if len(job["logs"]) > 300:
        job["logs"] = job["logs"][-300:]
    print(msg)


def run_job(job_id: str):
    """后台执行下载任务"""
    job = _JOBS.get(job_id)
    if not job:
        return
    job["status"] = "scanning"
    _log(job, f"开始下载知乎专栏: {job['columnId']}")
    try:
        info = fetch_column_info(job["columnId"])
        job["columnName"] = info["title"]
        _log(job, f"专栏名称: {info['title']}（作者：{info.get('author') or '未知'}）")

        items = fetch_column_items(
            job["columnId"],
            max_items=job.get("maxItems") or 0,
            log=lambda msg: _log(job, msg),
        )
        normalized = [_extract_item(i) for i in items]
        job["total"] = len(normalized)
        job["items"] = [
            {k: v for k, v in n.items() if k not in ("content", "raw", "video")}
            for n in normalized
        ]
        if not normalized:
            raise RuntimeError("该专栏下没有可下载的内容")
        _log(job, f"共发现 {len(normalized)} 条内容（文章/回答/视频）")

        out_dir = Path(job["outputDir"]) / f"{_slugify(info['title'])}_{job['columnId']}"
        md_dir = out_dir / "markdown"
        video_dir = out_dir / "videos"
        md_dir.mkdir(parents=True, exist_ok=True)
        job["outputDir"] = str(out_dir)

        rows = []
        failed = 0
        job["status"] = "downloading"
        for idx, item in enumerate(normalized, 1):
            job["progress"] = idx - 1
            job["currentTitle"] = item["title"]
            _log(job, f"[{idx}/{len(normalized)}] {item['type']}: {item['title'][:50]}")

            entry = {k: v for k, v in item.items() if k not in ("content", "raw", "video")}
            try:
                html = _fetch_item_content(item)
                content_md = html_to_markdown(html, item["url"])
                if not content_md:
                    raise RuntimeError("正文解析为空")
                md = build_markdown(item, content_md)
                filename = f"{idx:04d}_{_slugify(item['title'], 50)}.md"
                (md_dir / filename).write_text(md, encoding="utf-8")
                entry["status"] = "ok"
                entry["markdownPath"] = str(md_dir / filename)

                # 视频：可选下载 mp4，否则仅在 Markdown 中保留链接
                if item["type"] in ("video", "zvideo"):
                    video_url = _find_video_url(item.get("video"))
                    if job.get("downloadVideos") and video_url:
                        vfile = video_dir / f"{idx:04d}_{_slugify(item['title'], 40)}.mp4"
                        ok = download_video(video_url, vfile)
                        entry["videoPath"] = str(vfile) if ok else ""
                        _log(job, f"      视频{'下载成功' if ok else '下载失败（已保留链接）'}")
            except Exception as e:
                failed += 1
                entry["status"] = "error"
                entry["error"] = str(e)[:200]
                _log(job, f"      处理失败: {e}")

            rows.append({**entry, "excerpt": item.get("excerpt", "")[:200]})
            job["progress"] = idx
            job["items"][idx - 1] = entry
            time.sleep(0.3)  # 适度限速，避免触发风控

        excel_path = out_dir / f"{_slugify(info['title'])}_{job['columnId']}_文章数据.xlsx"
        generate_excel(rows, excel_path)
        job["excelPath"] = str(excel_path)

        # ============ 自动导入知识库 ============
        import_result = {"imported": 0, "skipped": 0, "failed": 0}
        if job.get("autoImport"):
            try:
                job["status"] = "importing"
                _log(job, "正在将下载内容导入知识库...")
                from kb_service import save_to_kb
                ok_items = [i for i in job["items"] if i.get("status") == "ok" and i.get("markdownPath")]
                for idx, entry in enumerate(ok_items, 1):
                    try:
                        md_text = Path(entry["markdownPath"]).read_text(encoding="utf-8")
                        result = save_to_kb(
                            url=entry.get("url", ""),
                            title=entry.get("title", ""),
                            original_md=md_text,
                            source="知乎专栏",
                            tags=["知乎"],
                            force=False,
                            build_index=False,
                        )
                        if result.get("success"):
                            import_result["imported"] += 1
                        else:
                            import_result["skipped"] += 1
                    except Exception as e:
                        import_result["failed"] += 1
                        _log(job, f"  导入失败 {entry.get('title', '')[:40]}: {e}")
                    job["progress"] = job["total"]
                    job["currentTitle"] = f"正在导入知识库 ({idx}/{len(ok_items)})"
                job["importResult"] = import_result
                _log(
                    job,
                    f"知识库导入完成：新增 {import_result['imported']}，"
                    f"跳过重复 {import_result['skipped']}，失败 {import_result['failed']}",
                )
                try:
                    from kb_index import ensure_kb_index
                    ensure_kb_index(force=True)
                    _log(job, "知识库索引已更新（作者 + 主题）")
                except Exception as e:
                    _log(job, f"知识库索引更新失败: {e}")
            except Exception as e:
                _log(job, f"知识库导入失败: {e}")

        job["status"] = "done"
        job["progress"] = job["total"]
        job["message"] = (
            f"完成：成功 {len(rows) - failed} 条，失败 {failed} 条"
            f"，Markdown 已保存至 {md_dir}，Excel：{excel_path}"
        )
        _log(job, f"✅ 下载完成，共 {len(rows)} 条（失败 {failed}）")
        _log(job, f"Excel: {excel_path}")
    except Exception as e:
        job["status"] = "error"
        job["error"] = str(e)[:500]
        job["message"] = f"下载失败: {e}"
        _log(job, f"❌ 下载失败: {e}")


def start_job(
    column_id: str,
    output_dir: str = "",
    download_videos: bool = False,
    max_items: int = 0,
    auto_import: bool = True,
) -> dict:
    """创建并启动后台下载任务"""
    col_id, resolved_from = resolve_column_input(column_id)
    job_id = uuid.uuid4().hex[:12]
    job = {
        "jobId": job_id,
        "columnId": col_id,
        "columnName": col_id,
        "resolvedFrom": resolved_from,
        "status": "queued",
        "progress": 0,
        "total": 0,
        "currentTitle": "",
        "message": "",
        "error": "",
        "createdAt": _now(),
        "outputDir": output_dir or str(DEFAULT_OUTPUT_DIR),
        "excelPath": "",
        "downloadVideos": bool(download_videos),
        "autoImport": bool(auto_import),
        "maxItems": int(max_items or 0),
        "importResult": None,
        "items": [],
        "logs": [],
    }
    with _JOBS_LOCK:
        _JOBS[job_id] = job
    threading.Thread(target=run_job, args=(job_id,), daemon=True).start()
    return job


def get_job(job_id: str) -> Optional[dict]:
    with _JOBS_LOCK:
        return _JOBS.get(job_id)


def list_jobs(limit: int = 10) -> list[dict]:
    with _JOBS_LOCK:
        jobs = sorted(_JOBS.values(), key=lambda j: j["createdAt"], reverse=True)
    return jobs[:limit]


# ==================== 直接运行（调试用） ====================

if __name__ == "__main__":
    _ensure_utf8_io()
    import sys

    sample = sys.argv[1] if len(sys.argv) > 1 else "c_1020247688083775488"
    print(f"测试专栏: {sample}")
    job = start_job(sample)
    print(f"jobId: {job['jobId']}")
    while True:
        j = get_job(job["jobId"])
        if j and j["status"] in ("done", "error"):
            print(f"\n状态: {j['status']}")
            print(j.get("message", ""))
            break
        time.sleep(2)
